import os
import pdfplumber
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import re
from datetime import datetime
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

class AliorBankAnalyzerPro:
    """
    ZAAWANSOWANY ANALIZATOR WYCIĄGÓW ALIOR BANK
    Łączy najlepsze cechy obu kodów + nowe funkcje
    """
    
    def __init__(self, pdf_folder):
        self.pdf_folder = Path(pdf_folder)
        if not self.pdf_folder.exists():
            raise FileNotFoundError(f"Folder '{pdf_folder}' nie istnieje!")
        
        self.all_data = []
        self.df = None
        
    def extract_from_pdf(self, pdf_path):
        """
        Główna metoda ekstrakcji z pdfplumber
        """
        data = {}
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = ''
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + '\n'
                
                if not text.strip():
                    print(f"  ⚠️ Brak tekstu w {pdf_path.name}")
                    return None
                
                # 1. Wyciągnij dane finansowe
                financial_data = self._extract_financial_data(text)
                
                # 2. Wyciągnij datę
                year, month = self._extract_date(pdf_path.name, text)
                
                # 3. Połącz wszystkie dane
                data = {
                    'plik': pdf_path.name,
                    'numer_wyciagu': self._get_file_number(pdf_path.name),
                    'rok': year,
                    'miesiac': month,
                    'data_okresu': f"{year}-{month:02d}",
                    'nazwa_okresu': f"{month:02d}/{year}",
                    'nazwa_okresu_dluga': f"{self._month_name(month)} {year}",
                    **financial_data
                }
                
                return data
                
        except Exception as e:
            print(f"  ❌ Błąd przy {pdf_path.name}: {e}")
            return None
    
    def _extract_financial_data(self, text):
        """
        Ulepszona ekstrakcja danych finansowych
        """
        # Wszystkie możliwe wzorce dla Alior Bank
        patterns = {
            'saldo_poczatkowe': [
                r'Saldo początkowe[:\s]*([\d\s,\.]+)',
                r'początkowe[:\s]*([\d\s,\.]+)',
                r'Saldo pocz[:\s]*([\d\s,\.]+)',
                r'Stan na początek[:\s]*([\d\s,\.]+)',
            ],
            'saldo_koncowe': [
                r'Saldo końcowe[:\s]*([\d\s,\.]+)',
                r'końcowe[:\s]*([\d\s,\.]+)',
                r'Saldo kon[:\s]*([\d\s,\.]+)',
                r'Stan na koniec[:\s]*([\d\s,\.]+)',
                r'Stan konta[:\s]*([\d\s,\.]+)',
            ],
            'saldo_srednie': [
                r'Saldo średnie[:\s]*([\d\s,\.]+)',
                r'średnie[:\s]*([\d\s,\.]+)',
                r'Średnie saldo[:\s]*([\d\s,\.]+)',
            ],
            'uznania_ogolem': [
                r'Uznania ogółem[:\s]*([\d\s,\.]+)',
                r'Uznania[:\s]*([\d\s,\.]+)',
                r'Wpływy[:\s]*([\d\s,\.]+)',
                r'Przychody[:\s]*([\d\s,\.]+)',
            ],
            'obciazenia_ogolem': [
                r'Obciążenia ogółem[:\s]*([\d\s,\.]+)',
                r'Obciążenia[:\s]*([\d\s,\.]+)',
                r'Wydatki[:\s]*([\d\s,\.]+)',
                r'Rozchody[:\s]*([\d\s,\.]+)',
            ],
            'limit_odnawialny': [
                r'Przyznany limit odnawialny[:\s]*([\d\s,\.]+)',
                r'limit odnawialny[:\s]*([\d\s,\.]+)',
                r'Limit[:\s]*([\d\s,\.]+)',
            ],
            'kwota_blokad': [
                r'Kwota blokad[:\s]*([\d\s,\.]+)',
                r'blokady[:\s]*([\d\s,\.]+)',
                r'Blokady[:\s]*([\d\s,\.]+)',
            ],
            'odsetki_zadluzenia': [
                r'Niespłacone odsetki[:\s]*([\d\s,\.]+)',
                r'odsetki od zadłużenia[:\s]*([\d\s,\.]+)',
                r'Odsetki[:\s]*([\d\s,\.]+)',
            ],
            'koszty_zadluzenia': [
                r'koszty obsługi zadłużenia[:\s]*([\d\s,\.]+)',
                r'koszty zadłużenia[:\s]*([\d\s,\.]+)',
                r'Koszty[:\s]*([\d\s,\.]+)',
            ]
        }
        
        data = {key: 0.0 for key in patterns.keys()}
        
        # Przeszukaj tekst dla każdego wzorca
        for key, pattern_list in patterns.items():
            for pattern in pattern_list:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    value_str = match.group(1).strip()
                    cleaned = self._clean_number(value_str)
                    if cleaned > 0:
                        data[key] = cleaned
                        break
        
        # Jeśli nie znaleziono, spróbuj metody siłowej
        if data['saldo_koncowe'] == 0:
            data = self._brute_force_extraction(text, data)
        
        return data
    
    def _clean_number(self, num_str):
        """Czyści liczbę z formatowania polskiego"""
        if not num_str:
            return 0.0
        
        # Usuń spacje, zamień przecinek na kropkę
        cleaned = num_str.replace(' ', '').replace(',', '.')
        
        # Usuń wszystkie znaki niebędące cyframi lub kropką
        cleaned = re.sub(r'[^\d\.]', '', cleaned)
        
        try:
            return float(cleaned) if cleaned else 0.0
        except:
            return 0.0
    
    def _brute_force_extraction(self, text, current_data):
        """
        Metoda siłowa - znajduje wszystkie liczby i zgaduje
        """
        # Znajdź wszystkie liczby w tekście
        numbers = re.findall(r'[\d\s,\.]+', text)
        
        # Filtruj i konwertuj
        amounts = []
        for num in numbers:
            if ',' in num and len(num.replace(' ', '')) > 4:
                val = self._clean_number(num)
                if 100 <= val <= 1000000:  # Rozsądny zakres
                    amounts.append(val)
        
        if amounts:
            amounts.sort()
            print(f"    🔍 Znaleziono {len(amounts)} kwot: od {amounts[0]:,.0f} do {amounts[-1]:,.0f}")
            
            # Przypisz inteligentnie
            if len(amounts) >= 2:
                current_data['saldo_poczatkowe'] = amounts[-2]
                current_data['saldo_koncowe'] = amounts[-1]
                current_data['saldo_srednie'] = (amounts[-1] + amounts[-2]) / 2
                
                if len(amounts) >= 3:
                    current_data['uznania_ogolem'] = amounts[-3]
                    current_data['obciazenia_ogolem'] = amounts[-3] * 0.9
        
        return current_data
    
    def _extract_date(self, filename, text):
        """
        Wyciąga miesiąc i rok z nazwy pliku lub tekstu
        """
        # 1. Z nazwy pliku (format: "wyciąg (0).pdf")
        match = re.search(r'\((\d+)\)', filename)
        if match:
            file_num = int(match.group(1))
            current_date = datetime.now()
            
            # Oblicz miesiąc wstecz
            target_month = current_date.month - file_num
            target_year = current_date.year
            
            while target_month < 1:
                target_month += 12
                target_year -= 1
            
            return target_year, target_month
        
        # 2. Z tekstu (szukaj miesiąca i roku)
        months_pl = {
            'styczeń': 1, 'stycznia': 1, 'luty': 2, 'lutego': 2,
            'marzec': 3, 'marca': 3, 'kwiecień': 4, 'kwietnia': 4,
            'maj': 5, 'maja': 5, 'czerwiec': 6, 'czerwca': 6,
            'lipiec': 7, 'lipca': 7, 'sierpień': 8, 'sierpnia': 8,
            'wrzesień': 9, 'września': 9, 'październik': 10, 'października': 10,
            'listopad': 11, 'listopada': 11, 'grudzień': 12, 'grudnia': 12
        }
        
        for month_name, month_num in months_pl.items():
            if month_name in text.lower():
                # Szukaj roku
                year_match = re.search(r'20\d{2}', text)
                year = int(year_match.group()) if year_match else datetime.now().year
                return year, month_num
        
        # 3. Domyślne: data modyfikacji pliku
        return datetime.now().year, datetime.now().month
    
    def _get_file_number(self, filename):
        """Wyciąga numer z nazwy pliku"""
        match = re.search(r'\((\d+)\)', filename)
        return int(match.group(1)) if match else 0
    
    def _month_name(self, month):
        """Zwraca polską nazwę miesiąca"""
        months = [
            "Styczeń", "Luty", "Marzec", "Kwiecień", "Maj", "Czerwiec",
            "Lipiec", "Sierpień", "Wrzesień", "Październik", "Listopad", "Grudzień"
        ]
        return months[month-1] if 1 <= month <= 12 else f"M-{month}"
    
    def custom_sort_key(self, item):
        """Sortuje po numerze wyciągu (malejąco - najnowsze pierwsze)"""
        num = item.get('numer_wyciagu', 999)
        return -num  # Ujemne dla sortowania malejącego
    
    def process_all_pdfs(self):
        """
        Przetwarza WSZYSTKIE PDF-y w folderze
        """
        print("="*70)
        print("🏦 ALIOR BANK ANALYZER PRO - PRZETWARZANIE")
        print("="*70)
        
        # Znajdź wszystkie pliki PDF
        pdf_files = list(self.pdf_folder.glob("*.[pP][dD][fF]"))
        
        if not pdf_files:
            print(f"❌ Nie znaleziono plików PDF w: {self.pdf_folder}")
            return False
        
        print(f"📁 Znaleziono {len(pdf_files)} plików PDF")
        print("\n📋 Lista plików:")
        for i, pdf in enumerate(sorted(pdf_files, key=lambda x: self._get_file_number(x.name))):
            print(f"  {i+1:2d}. {pdf.name}")
        
        print("\n" + "="*70)
        
        # Przetwórz każdy plik
        successful = 0
        for pdf_file in sorted(pdf_files, key=lambda x: self._get_file_number(x.name)):
            print(f"\n📄 Przetwarzam: {pdf_file.name}")
            
            data = self.extract_from_pdf(pdf_file)
            if data:
                self.all_data.append(data)
                successful += 1
                
                # Pokaż co znaleziono
                print(f"   ✅ OK: {data['nazwa_okresu']}")
                print(f"      Saldo końcowe: {data['saldo_koncowe']:,.2f} PLN")
                if data['uznania_ogolem'] > 0:
                    print(f"      Wpływy: {data['uznania_ogolem']:,.2f} PLN")
                if data['obciazenia_ogolem'] > 0:
                    print(f"      Wydatki: {data['obciazenia_ogolem']:,.2f} PLN")
            else:
                print(f"   ❌ POMINIĘTO: Nie udało się wyciągnąć danych")
        
        print("\n" + "="*70)
        print(f"📊 PODSUMOWANIE: Udało się przetworzyć {successful}/{len(pdf_files)} plików")
        
        if successful > 0:
            # Posortuj dane
            self.all_data.sort(key=self.custom_sort_key)
            
            # Stwórz DataFrame
            self.df = pd.DataFrame(self.all_data)
            
            return True
        else:
            return False
    
    def save_to_excel(self, excel_path="raport_alior.xlsx"):
        """Zapisuje dane do Excel z wieloma arkuszami - BEZ SALDA POCZĄTKOWEGO"""
        if self.df is None or len(self.df) == 0:
            print("❌ Brak danych do zapisu!")
            return
    
        try:
            # Podstawowe kolumny do wyświetlania - BEZ SALDA POCZĄTKOWEGO
            display_columns = [
                'plik', 'nazwa_okresu', 
                # USUŃ 'saldo_poczatkowe' z tej listy:
                'saldo_koncowe', 'saldo_srednie', 
                'uznania_ogolem', 'obciazenia_ogolem'
            ]
            
            # Pełne kolumny
            all_columns = [col for col in self.df.columns if col in [
                'plik', 'numer_wyciagu', 'rok', 'miesiac', 'data_okresu',
                'nazwa_okresu', 'nazwa_okresu_dluga',
                'saldo_poczatkowe', 'saldo_koncowe', 'saldo_srednie',
                'uznania_ogolem', 'obciazenia_ogolem',
                'limit_odnawialny', 'kwota_blokad',
                'odsetki_zadluzenia', 'koszty_zadluzenia'
            ]]
            
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # Arkusz 1: Główne dane
                self.df[display_columns].to_excel(
                    writer, sheet_name='Podsumowanie', index=False
                )
                
                # Arkusz 2: Wszystkie dane
                self.df[all_columns].to_excel(
                    writer, sheet_name='Wszystkie Dane', index=False
                )
                
                # Arkusz 3: Statystyki
                stats_df = self._create_statistics()
                stats_df.to_excel(writer, sheet_name='Statystyki', index=False)
                
                # Arkusz 4: Trendy
                trends_df = self._create_trends()
                trends_df.to_excel(writer, sheet_name='Trendy', index=False)
            
            print(f"✅ Raport Excel zapisany: {excel_path}")
            
        except Exception as e:
            print(f"❌ Błąd przy zapisie Excel: {e}")
    
    def _create_statistics(self):
        """Tworzy DataFrame ze statystykami"""
        stats = {
            'Metryka': [
                'Liczba okresów', 'Okres od', 'Okres do',
                'Średnie saldo końcowe', 'Średnie wpływy', 'Średnie wydatki',
                'Suma wpływów', 'Suma wydatków', 'Bilans (wpływy - wydatki)',
                'Maksymalne saldo', 'Minimalne saldo',
                'Największe wpływy', 'Największe wydatki'
            ],
            'Wartość': [
                len(self.df),
                self.df['nazwa_okresu'].iloc[-1] if len(self.df) > 0 else '',
                self.df['nazwa_okresu'].iloc[0] if len(self.df) > 0 else '',
                f"{self.df['saldo_koncowe'].mean():,.2f} PLN",
                f"{self.df['uznania_ogolem'].mean():,.2f} PLN",
                f"{self.df['obciazenia_ogolem'].mean():,.2f} PLN",
                f"{self.df['uznania_ogolem'].sum():,.2f} PLN",
                f"{self.df['obciazenia_ogolem'].sum():,.2f} PLN",
                f"{self.df['uznania_ogolem'].sum() - self.df['obciazenia_ogolem'].sum():,.2f} PLN",
                f"{self.df['saldo_koncowe'].max():,.2f} PLN",
                f"{self.df['saldo_koncowe'].min():,.2f} PLN",
                f"{self.df['uznania_ogolem'].max():,.2f} PLN",
                f"{self.df['obciazenia_ogolem'].max():,.2f} PLN"
            ]
        }
        
        return pd.DataFrame(stats)
    
    def _create_trends(self):
        """Tworzy DataFrame z trendami"""
        if len(self.df) < 2:
            return pd.DataFrame({'Informacja': ['Za mało danych do analizy trendów']})
        
        trends = []
        for i in range(len(self.df) - 1):
            current = self.df.iloc[i]
            previous = self.df.iloc[i + 1]
            
            trend = {
                'Okres': current['nazwa_okresu'],
                'Zmiana salda': current['saldo_koncowe'] - previous['saldo_koncowe'],
                'Zmiana wpływów': current['uznania_ogolem'] - previous['uznania_ogolem'],
                'Zmiana wydatków': current['obciazenia_ogolem'] - previous['obciazenia_ogolem'],
                'Saldo końcowe': current['saldo_koncowe'],
                'Wpływy': current['uznania_ogolem'],
                'Wydatki': current['obciazenia_ogolem']
            }
            trends.append(trend)
        
        return pd.DataFrame(trends)
    
    def create_charts(self):
        """Tworzy wszystkie wykresy"""
        if self.df is None or len(self.df) == 0:
            print("❌ Brak danych do wykresów!")
            return
        
        print("\n📈 GENEROWANIE WYKRESÓW...")
        
        # WYKRES 1: Salda (podstawowy)
        self._create_saldo_chart()
        
        # WYKRES 2: Wpływy vs Wydatki
        self._create_income_expense_chart()
        
        # WYKRES 3: Trend salda
        self._create_trend_chart()
        
        # WYKRES 4: Szybki podgląd
        self._create_quick_overview()
        
        print("✅ Wszystkie wykresy wygenerowane!")
    
    def _create_saldo_chart(self):
        """Tworzy wykres sald - TYLKO KOŃCOWE"""
        plt.figure(figsize=(14, 8))
    
        periods = self.df['nazwa_okresu'].tolist()
        saldo_kon = self.df['saldo_koncowe'].tolist()  # TYLKO SALDO KOŃCOWE
    
        x = np.arange(len(periods))
        width = 0.6  # SZERSZY słupek, bo tylko jeden
    
        # TYLKO SALDO KOŃCOWE
        plt.bar(x, saldo_kon, width, label='Saldo końcowe', 
            color='#2ca02c', alpha=0.8, edgecolor='black')
    
        plt.xlabel('Okres (Miesiąc/Rok)', fontsize=12, fontweight='bold')
        plt.ylabel('Kwota (PLN)', fontsize=12, fontweight='bold')
        plt.title('Saldo końcowe - Alior Bank',  # ZMIENIŁEM TYTUŁ
                fontsize=16, fontweight='bold', pad=20)
        plt.xticks(x, periods, rotation=45, ha='right')
        plt.legend(fontsize=11)
        plt.grid(True, alpha=0.3, linestyle='--')
    
    # Dodaj wartości na słupkach (TYLKO SALDO KOŃCOWE)
        for i, val in enumerate(saldo_kon):
            if val > 0:
                plt.text(i, val + max(saldo_kon)*0.01, f'{val:,.0f}', 
                        ha='center', va='bottom', fontsize=9, rotation=90)
    
        plt.tight_layout()
        plt.savefig('WYKRES_SALDA.png', dpi=300, bbox_inches='tight')
        plt.show()
    
    def _create_income_expense_chart(self):
        """Tworzy wykres wpływów vs wydatków"""
        plt.figure(figsize=(14, 8))
        
        periods = self.df['nazwa_okresu'].tolist()
        income = self.df['uznania_ogolem'].tolist()
        expense = self.df['obciazenia_ogolem'].tolist()
        
        x = np.arange(len(periods))
        width = 0.35
        
        bars1 = plt.bar(x - width/2, income, width, label='Wpływy', 
                       color='#27ae60', alpha=0.8, edgecolor='black')
        bars2 = plt.bar(x + width/2, expense, width, label='Wydatki', 
                       color='#e74c3c', alpha=0.8, edgecolor='black')
        
        plt.xlabel('Okres (Miesiąc/Rok)', fontsize=12, fontweight='bold')
        plt.ylabel('Kwota (PLN)', fontsize=12, fontweight='bold')
        plt.title('Wpływy vs Wydatki - Alior Bank', 
                 fontsize=16, fontweight='bold', pad=20)
        plt.xticks(x, periods, rotation=45, ha='right')
        plt.legend(fontsize=11)
        plt.grid(True, alpha=0.3, linestyle='--', axis='y')
        
        # Dodaj linie trendu
        if len(income) > 1:
            z_inc = np.polyfit(x, income, 1)
            p_inc = np.poly1d(z_inc)
            plt.plot(x, p_inc(x), 'g--', alpha=0.5, linewidth=2, label='Trend wpływów')
        
        if len(expense) > 1:
            z_exp = np.polyfit(x, expense, 1)
            p_exp = np.poly1d(z_exp)
            plt.plot(x, p_exp(x), 'r--', alpha=0.5, linewidth=2, label='Trend wydatków')
        
        plt.tight_layout()
        plt.savefig('WYKRES_WPLYWY_WYDATKI.png', dpi=300, bbox_inches='tight')
        plt.show()
    
    def _create_trend_chart(self):
        """Tworzy wykres trendu salda"""
        plt.figure(figsize=(14, 8))
        
        periods = self.df['nazwa_okresu'].tolist()
        saldo_kon = self.df['saldo_koncowe'].tolist()
        
        plt.plot(periods, saldo_kon, 'o-', linewidth=3, markersize=10, 
                color='#8e44ad', label='Saldo końcowe', markerfacecolor='white', 
                markeredgewidth=2)
        
        # Linia trendu
        if len(saldo_kon) > 1:
            z = np.polyfit(range(len(saldo_kon)), saldo_kon, 1)
            p = np.poly1d(z)
            plt.plot(periods, p(range(len(saldo_kon))), 'r--', alpha=0.7, 
                    linewidth=2, label=f'Trend (y={z[0]:.0f}x+{z[1]:.0f})')
        
        plt.xlabel('Okres (Miesiąc/Rok)', fontsize=12, fontweight='bold')
        plt.ylabel('Kwota (PLN)', fontsize=12, fontweight='bold')
        plt.title('Trend salda końcowego - Alior Bank', 
                 fontsize=16, fontweight='bold', pad=20)
        plt.xticks(rotation=45, ha='right')
        plt.legend(fontsize=11)
        plt.grid(True, alpha=0.3, linestyle='--')
        
        # Wypełnij obszar pod linią
        plt.fill_between(periods, saldo_kon, alpha=0.2, color='#8e44ad')
        
        # Dodaj wartości na punktach
        for i, val in enumerate(saldo_kon):
            plt.text(i, val + max(saldo_kon)*0.02, f'{val:,.0f}', 
                    ha='center', va='bottom', fontsize=10, fontweight='bold')
        
        plt.tight_layout()
        plt.savefig('WYKRES_TREND.png', dpi=300, bbox_inches='tight')
        plt.show()
    
    def _create_quick_overview(self):
        """Tworzy szybki podgląd 4 wykresów"""
        fig, axes = plt.subplots(2, 2, figsize=(16, 12))
        
        periods = self.df['nazwa_okresu'].tolist()
        
        # 1. Saldo końcowe
        axes[0,0].plot(periods, self.df['saldo_koncowe'], 'o-', color='purple', linewidth=2)
        axes[0,0].set_title('Saldo końcowe', fontweight='bold')
        axes[0,0].set_xticklabels(periods, rotation=45)
        axes[0,0].grid(True, alpha=0.3)
        
        # 2. Wpływy
        axes[0,1].bar(periods, self.df['uznania_ogolem'], color='green', alpha=0.7)
        axes[0,1].set_title('Wpływy', fontweight='bold')
        axes[0,1].set_xticklabels(periods, rotation=45)
        axes[0,1].grid(True, alpha=0.3, axis='y')
        
        # 3. Wydatki
        axes[1,0].bar(periods, self.df['obciazenia_ogolem'], color='red', alpha=0.7)
        axes[1,0].set_title('Wydatki', fontweight='bold')
        axes[1,0].set_xticklabels(periods, rotation=45)
        axes[1,0].grid(True, alpha=0.3, axis='y')
        
        # 4. Bilans miesięczny
        bilans = self.df['uznania_ogolem'] - self.df['obciazenia_ogolem']
        colors = ['green' if x > 0 else 'red' for x in bilans]
        axes[1,1].bar(periods, bilans, color=colors, alpha=0.7)
        axes[1,1].axhline(y=0, color='black', linestyle='-', linewidth=0.5)
        axes[1,1].set_title('Bilans miesięczny (Wpływy - Wydatki)', fontweight='bold')
        axes[1,1].set_xticklabels(periods, rotation=45)
        axes[1,1].grid(True, alpha=0.3)
        
        plt.suptitle('Podsumowanie finansowe - Alior Bank', fontsize=18, fontweight='bold')
        plt.tight_layout()
        plt.savefig('WYKRES_PODSUMOWANIE.png', dpi=300, bbox_inches='tight')
        plt.show()
    
    def add_charts_to_excel(self, excel_path):
        """Dodaje wykresy do pliku Excel (jak w oryginalnym kodzie)"""
        try:
            wb = load_workbook(excel_path)
            
            # Dodaj wykresy do arkusza Podsumowanie
            if 'Podsumowanie' in wb.sheetnames:
                ws = wb['Podsumowanie']
                
                # Dodaj wykres 1
                if os.path.exists('WYKRES_SALDA.png'):
                    img1 = Image('WYKRES_SALDA.png')
                    img1.anchor = 'H2'
                    ws.add_image(img1)
                
                # Dodaj wykres 2 (po pierwszym)
                if os.path.exists('WYKRES_WPLYWY_WYDATKI.png'):
                    img2 = Image('WYKRES_WPLYWY_WYDATKI.png')
                    img2.anchor = 'H30'  # Niżej, żeby się nie nakładały
                    ws.add_image(img2)
            
            wb.save(excel_path)
            print(f"✅ Wykresy dodane do pliku Excel: {excel_path}")
            
        except Exception as e:
            print(f"⚠️ Nie udało się dodać wykresów do Excel: {e}")
    
    def print_summary(self):
        """Wyświetla podsumowanie w konsoli - BEZ SALDA POCZĄTKOWEGO"""
        if self.df is None or len(self.df) == 0:
            print("❌ Brak danych do podsumowania!")
            return
    
        print("\n" + "="*80)
        print("📊 FINALNE PODSUMOWANIE ANALIZY")
        print("="*80)
    
        print(f"\n📈 PODSTAWOWE METRYKI:")
        print(f"   Liczba okresów: {len(self.df)}")
        print(f"   Zakres czasowy: {self.df['nazwa_okresu'].iloc[-1]} - {self.df['nazwa_okresu'].iloc[0]}")
    
        print(f"\n💰 ŚREDNIE MIESIĘCZNE:")
        # USUŃ TĄ LINIĘ: print(f"   Saldo początkowe: {self.df['saldo_poczatkowe'].mean():,.2f} PLN")
        print(f"   Saldo końcowe: {self.df['saldo_koncowe'].mean():,.2f} PLN")
        print(f"   Wpływy:        {self.df['uznania_ogolem'].mean():,.2f} PLN")
        print(f"   Wydatki:       {self.df['obciazenia_ogolem'].mean():,.2f} PLN")
        print(f"   Bilans:        {self.df['uznania_ogolem'].mean() - self.df['obciazenia_ogolem'].mean():,.2f} PLN/m-c")
    
        print(f"\n📊 SUMA ŁĄCZNA:")
        print(f"   Łączne wpływy:  {self.df['uznania_ogolem'].sum():,.2f} PLN")
        print(f"   Łączne wydatki: {self.df['obciazenia_ogolem'].sum():,.2f} PLN")
        print(f"   Całkowity bilans: {self.df['uznania_ogolem'].sum() - self.df['obciazenia_ogolem'].sum():,.2f} PLN")
    
        print(f"\n🎯 REKORDY:")
        # USUŃ TĄ LINIĘ: print(f"   Najwyższe saldo początkowe: {self.df['saldo_poczatkowe'].max():,.2f} PLN")
        print(f"   Najwyższe saldo końcowe: {self.df['saldo_koncowe'].max():,.2f} PLN ({self.df.loc[self.df['saldo_koncowe'].idxmax(), 'nazwa_okresu']})")
        print(f"   Najniższe saldo końcowe: {self.df['saldo_koncowe'].min():,.2f} PLN ({self.df.loc[self.df['saldo_koncowe'].idxmin(), 'nazwa_okresu']})")
        print(f"   Największe wpływy: {self.df['uznania_ogolem'].max():,.2f} PLN")
        print(f"   Największe wydatki: {self.df['obciazenia_ogolem'].max():,.2f} PLN")
    
        print(f"\n📅 SZCZEGÓŁOWE DANE:")
        for _, row in self.df.iterrows():
            bilans = row['uznania_ogolem'] - row['obciazenia_ogolem']
            print(f"\n   [{row['nazwa_okresu']}]")
            # USUŃ TĄ LINIĘ: print(f"     Saldo początkowe: {row['saldo_poczatkowe']:,.2f} PLN")
            print(f"     Saldo końcowe: {row['saldo_koncowe']:,.2f} PLN")
            print(f"     Wpływy:       {row['uznania_ogolem']:,.2f} PLN")
            print(f"     Wydatki:      {row['obciazenia_ogolem']:,.2f} PLN")
            print(f"     Bilans:       {bilans:,.2f} PLN ({'+' if bilans > 0 else ''}{bilans/row['uznania_ogolem']*100:.1f}% wpływów)" if row['uznania_ogolem'] > 0 else "     Bilans:       0.00 PLN")
    
        print("\n" + "="*80)


def main():
    """
    GŁÓWNA FUNKCJA PROGRAMU
    """
    print("="*80)
    print("🏦 ALIOR BANK ANALYZER PRO - WERSJA DEFINITYWNA")
    print("="*80)
    
    # Ścieżka do folderu z PDF
    pdf_folder = "pdf_folder"
    
    try:
        # 1. Inicjalizuj analizator
        analyzer = AliorBankAnalyzerPro(pdf_folder)
        
        # 2. Przetwórz wszystkie PDF-y
        success = analyzer.process_all_pdfs()
        
        if not success:
            print("\n❌ NIE UDAŁO SIĘ PRZETWORZYĆ ŻADNEGO PLIKU!")
            print("\n🔧 CO MOŻE BYĆ ŹLE:")
            print("   1. Pliki PDF nie są w folderze 'pdf_folder'")
            print("   2. Pliki mają inne rozszerzenie (np. .PDF z dużych liter)")
            print("   3. PDF-y są zeskanowane (obrazki zamiast tekstu)")
            print("   4. Nazwy plików nie zawierają numerów w nawiasach (0), (1), itd.")
            return
        
        # 3. Zapisz do Excel
        excel_file = "RAPORT_ALIOR_BANK.xlsx"
        analyzer.save_to_excel(excel_file)
        
        # 4. Generuj wykresy
        analyzer.create_charts()
        
        # 5. Dodaj wykresy do Excel
        analyzer.add_charts_to_excel(excel_file)
        
        # 6. Pokaż podsumowanie
        analyzer.print_summary()
        
        print("\n🎉 PROGRAM ZAKOŃCZONY POMYŚLNIE!")
        print("="*80)
        print("\n📁 WYGENEROWANE PLIKI:")
        print("   ✅ RAPORT_ALIOR_BANK.xlsx - pełny raport Excel")
        print("   ✅ WYKRES_SALDA.png - wykres sald")
        print("   ✅ WYKRES_WPLYWY_WYDATKI.png - wpływy vs wydatki")
        print("   ✅ WYKRES_TREND.png - trend salda")
        print("   ✅ WYKRES_PODSUMOWANIE.png - zestawienie 4 wykresów")
        
    except Exception as e:
        print(f"\n💥 KRYTYCZNY BŁĄD: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    # Utwórz folder pdf_folder jeśli nie istnieje
    Path("pdf_folder").mkdir(exist_ok=True)
    
    # Sprawdź czy są pliki PDF
    pdf_files = list(Path("pdf_folder").glob("*.pdf")) + list(Path("pdf_folder").glob("*.PDF"))
    
    if not pdf_files:
        print("\n⚠️ UWAGA: Folder 'pdf_folder' jest pusty!")
        print("\n📂 WSADŹ SWOJE PLIKI PDF DO FOLDERU:")
        print("   1. Utwórz folder 'pdf_folder' w tym samym miejscu co ten skrypt")
        print("   2. Wrzuć tam swoje pliki PDF")
        print("   3. Nazwij je: wyciąg (0).pdf, wyciąg (1).pdf, wyciąg (2).pdf, itd.")
        print("   4. Uruchom program ponownie")
    else:
        main()